from __future__ import annotations

import os
import tempfile
from collections import Counter
from pathlib import Path
from typing import List, Union

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import ConversionResult, VoterRecord

BLUE = "2F75B5"
LIGHT_BLUE = "D9EAF7"
VERY_LIGHT_BLUE = "EDF5FB"
WHITE = "FFFFFF"
DARK = "1F2937"
ORANGE = "FCE4D6"
RED = "F4CCCC"
FONT_NAME = "Nirmala UI"

HEADERS = [
    "क्रम संख्या",
    "EPIC संख्या",
    "नाम",
    "संबंध",
    "पिता/पति का नाम",
    "मकान संख्या",
    "आयु",
    "लिंग",
    "फोटो उपलब्ध",
    "विधानसभा क्षेत्र",
    "अनुभाग",
    "भाग संख्या",
    "स्रोत पृष्ठ",
]


def _border() -> Border:
    side = Side(style="thin", color="7F8C8D")
    return Border(left=side, right=side, top=side, bottom=side)


def _style(cell, fill=None, bold=False, color=DARK, size=10, align="left") -> None:
    cell.font = Font(name=FONT_NAME, bold=bold, color=color, size=size)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = _border()


def _record_values(record: VoterRecord) -> List[Union[str, int]]:
    return [
        record.serial_number,
        record.epic_id,
        record.name,
        record.relation,
        record.related_person_name,
        record.house_number,
        record.age,
        record.gender,
        record.photo_available,
        record.constituency,
        record.section,
        record.part_number,
        record.source_page,
    ]


def _metadata_rows(sheet, result: ConversionResult) -> None:
    blocks = [
        (1, 2, 2, 5, "विधानसभा निर्वाचन क्षेत्र", result.metadata.get("constituency", "")),
        (6, 2, 7, 9, "अनुभाग संख्या और नाम", result.metadata.get("section", "")),
        (10, 2, 11, 11, "भाग संख्या", result.metadata.get("part_number", "")),
        (12, 2, 13, 13, "कुल अभिलेख", len(result.records)),
    ]
    for label_col, row, value_start, value_end, label, value in blocks:
        sheet.cell(row=row, column=label_col, value=label)
        if value_end > value_start:
            sheet.merge_cells(
                start_row=row, start_column=value_start, end_row=row, end_column=value_end
            )
        sheet.cell(row=row, column=value_start, value=value)
        _style(sheet.cell(row=row, column=label_col), fill=LIGHT_BLUE, bold=True)
        for col in range(value_start, value_end + 1):
            _style(sheet.cell(row=row, column=col), fill=WHITE, align="center")
    sheet.cell(row=3, column=1, value="स्रोत")
    sheet.merge_cells(start_row=3, start_column=2, end_row=3, end_column=13)
    sheet.cell(row=3, column=2, value="अपलोड की गई चुनाव सूची की PDF")
    _style(sheet.cell(row=3, column=1), fill=LIGHT_BLUE, bold=True)
    for col in range(2, 14):
        _style(sheet.cell(row=3, column=col), fill=WHITE)


def write_excel(result: ConversionResult, output_path: Union[str, Path]) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "मतदाता डेटा"
    sheet.sheet_view.showGridLines = False

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=13)
    title = sheet.cell(row=1, column=1, value="निर्वाचक नामावली – मतदाता विवरण")
    title.font = Font(name=FONT_NAME, bold=True, color=WHITE, size=15)
    title.fill = PatternFill("solid", fgColor=BLUE)
    title.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 30

    _metadata_rows(sheet, result)
    header_row = 5
    for column, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=header_row, column=column, value=header)
        _style(cell, fill=BLUE, bold=True, color=WHITE, size=10, align="center")
    sheet.row_dimensions[header_row].height = 34

    for row_number, record in enumerate(result.records, start=header_row + 1):
        for column, value in enumerate(_record_values(record), start=1):
            cell = sheet.cell(row=row_number, column=column, value=value)
            fill = ORANGE if record.review_required else (VERY_LIGHT_BLUE if row_number % 2 else WHITE)
            _style(
                cell,
                fill=fill,
                align="center" if column in (1, 2, 4, 7, 8, 9, 12, 13) else "left",
            )
            if column in (1, 2, 6, 7, 12):
                cell.number_format = "@"
        sheet.row_dimensions[row_number].height = 24

    last_row = header_row + len(result.records)
    sheet.auto_filter.ref = "A{}:M{}".format(header_row, last_row)
    sheet.freeze_panes = "A6"
    widths = [12, 17, 21, 12, 24, 15, 9, 12, 15, 21, 18, 12, 12]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.print_title_rows = "1:5"

    summary = workbook.create_sheet("सारांश")
    summary.sheet_view.showGridLines = False
    summary.merge_cells("A1:D1")
    summary["A1"] = "रूपांतरण सारांश"
    summary["A1"].font = Font(name=FONT_NAME, bold=True, color=WHITE, size=15)
    summary["A1"].fill = PatternFill("solid", fgColor=BLUE)
    summary["A1"].alignment = Alignment(horizontal="center")
    gender_counts = Counter(record.gender or "अज्ञात" for record in result.records)
    epic_counts = Counter(record.epic_id for record in result.records if record.epic_id)
    duplicate_count = sum(count - 1 for count in epic_counts.values() if count > 1)
    rows = [
        ("कुल PDF पृष्ठ", result.page_count),
        ("पहचाने गए कार्ड स्थान", result.detected_card_count),
        ("कुल निर्यातित अभिलेख", len(result.records)),
        ("पुरुष", gender_counts.get("पुरुष", 0)),
        ("महिला", gender_counts.get("महिला", 0)),
        ("अन्य/अज्ञात", sum(v for k, v in gender_counts.items() if k not in ("पुरुष", "महिला"))),
        ("समीक्षा आवश्यक", result.review_count),
        ("डुप्लिकेट EPIC", duplicate_count),
        ("प्रक्रिया समय (सेकंड)", round(result.elapsed_seconds, 2)),
    ]
    for row_index, (label, value) in enumerate(rows, start=3):
        summary.cell(row=row_index, column=1, value=label)
        summary.cell(row=row_index, column=2, value=value)
        _style(summary.cell(row=row_index, column=1), fill=LIGHT_BLUE, bold=True)
        _style(summary.cell(row=row_index, column=2), fill=WHITE, align="center")
    summary.column_dimensions["A"].width = 30
    summary.column_dimensions["B"].width = 24

    review = workbook.create_sheet("समीक्षा")
    review_headers = [
        "क्रम संख्या",
        "EPIC संख्या",
        "नाम",
        "स्रोत पृष्ठ",
        "कार्ड",
        "विश्वास %",
        "कारण",
        "मूल OCR पाठ",
    ]
    for col, header in enumerate(review_headers, start=1):
        _style(review.cell(row=1, column=col, value=header), fill=BLUE, bold=True, color=WHITE, align="center")
    review_records = [record for record in result.records if record.review_required or record.duplicate_epic]
    for row_index, record in enumerate(review_records, start=2):
        values = [
            record.serial_number,
            record.epic_id,
            record.name,
            record.source_page,
            record.source_card,
            round(record.confidence * 100, 1),
            record.review_reason,
            record.raw_text,
        ]
        for col, value in enumerate(values, start=1):
            _style(review.cell(row=row_index, column=col, value=value), fill=ORANGE, align="left")
    review.freeze_panes = "A2"
    review.auto_filter.ref = "A1:H{}".format(max(1, len(review_records) + 1))
    review.sheet_view.showGridLines = False
    for col, width in enumerate([12, 17, 22, 12, 10, 12, 38, 75], start=1):
        review.column_dimensions[get_column_letter(col)].width = width

    log = workbook.create_sheet("प्रक्रिया लॉग")
    _style(log.cell(row=1, column=1, value="चेतावनी / प्रक्रिया नोट"), fill=BLUE, bold=True, color=WHITE)
    for row_index, warning in enumerate(result.warnings or ["कोई प्रक्रिया चेतावनी नहीं।"], start=2):
        _style(log.cell(row=row_index, column=1, value=warning), fill=WHITE)
    log.column_dimensions["A"].width = 110
    log.sheet_view.showGridLines = False

    # Save to a temporary file and reopen it before publishing. This catches
    # structural corruption and prevents Excel from showing "[Repaired]".
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False, dir=str(output_path.parent)) as handle:
        temporary_path = Path(handle.name)
    try:
        workbook.save(str(temporary_path))
        check = load_workbook(str(temporary_path), read_only=False, data_only=False)
        check.close()
        os.replace(str(temporary_path), str(output_path))
    finally:
        if temporary_path.exists():
            temporary_path.unlink()
    return output_path
